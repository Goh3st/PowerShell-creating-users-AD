# Import required modules
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
import time
import openpyxl
from openpyxl import Workbook
headlines = ['Street' , 'Area' , 'Number of rooms' , 'Floor Number' , 'Square meteres' , 'Price' , 'Walk distance from the University' , 'More Information' , 'The ad URL']

#prerequimentes

options = webdriver.ChromeOptions()
options.add_argument("user-data-dir=C:\\Users\\גלעד\\AppData\\Local\\Google\\Chrome\\User Data")  #Path to your chrome profile
w = webdriver.Chrome(executable_path=r"C:\Users\גלעד\Documents\Projects\MusicDownloader\chromedriver.exe", options=options)

#fetching data function
def datafetch():
    try:
        book = openpyxl.load_workbook('Apartments.xlsx')
    except:
        book = Workbook()
        sheetB = book.active
        sheetB.title = 'שכונה ב'
        sheetG = book.create_sheet(title = 'שכונה ג' )
        sheetD = book.create_sheet(title='שכונה ד')
        sheetW = book.create_sheet(title  = 'שכונה ו' )
        sheetB.append(headlines)
        sheetG.append(headlines)
        sheetD.append(headlines)
        sheetW.append(headlines)
    sheetB = book.get_sheet_by_name('שכונה ב')
    sheetG = book.get_sheet_by_name('שכונה ג')
    sheetD = book.get_sheet_by_name('שכונה ד')
    sheetW = book.get_sheet_by_name('שכונה ו')
    url = input('Enter Url Of The Apartment Information: ')
    w.get(url)
    time.sleep(10)
    data = w.find_element_by_xpath('//*[@id="__layout"]/div/main/div/div[3]/div[5]/div/div[1]/div/div/div[2]/div[2]')
    moredescription = w.find_element_by_xpath('/html/body/div[2]/div[2]/div/main/div/div[3]/div[5]/div/div[2]/div[2]/div[1]/section[1]/div[1]/div/div/div/p')
    print(data.text)
    y = data.text.splitlines()
    y.pop(3)
    y.pop(4)
    y.pop(5)
    print(y)
    y[4] = y[4].replace("₪", "")
    walktime = input('Check walk distance: ')
    y.append(walktime)
    y.append(moredescription.text)
    y.append(url)
    if(sheetB.title in y[1]):
        sheetB.append(y)
    if(sheetG.title in y[1]):
        sheetG.append(y)
    if(sheetD.title in y[1]):
        sheetD.append(y)
    if(sheetW.title in y[1]):
        sheetW.append(y)
    book.save('Apartments.xlsx')
    datafetch()

datafetch()
