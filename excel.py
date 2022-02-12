import openpyxl
from openpyxl import Workbook
headlines = ['Street' , 'Area' , 'Number of rooms' , 'Floor Number' , 'Square meteres' , 'Price' , 'Walk distance from the University' , 'More Information' , 'The ad URL']
try:
    book = openpyxl.load_workbook('appending.xlsx')
except:
    book = Workbook()
    sheetB = book.active
    sheetB.title = 'שכונה ב'
    sheetG = book.create_sheet(title = 'שכונה ג' )
    sheetD = book.create_sheet(title='שכונה ד')
    sheetW = book.create_sheet(title  = 'שכונה ו' )
    sheetB.append(headlines)
    sheetG.append(headlines)
    sheetD.append(headlines)
    sheetW.append(headlines)
sheetB = book.get_sheet_by_name('שכונה ב')
sheetG = book.get_sheet_by_name('שכונה ג')
sheetD = book.get_sheet_by_name('שכונה ד')
sheetW = book.get_sheet_by_name('שכונה ו')
myList = ['First','Second','Third','Fourth','Fifth']
#sheet['A1'] = 'רחוב' #y[0]
#sheet['B1'] = 'אזור' #y[1]
#sheet['C1'] = 'חדרים' #y[2]
#sheet['D1'] = 'קומה' #y[4]
#sheet['E1'] = 'מטר רבוע' #y[6]
#sheet['F1'] = 'מחיר' #y[8]
#sheet['G1'] = 'מרחק הליכה מהאוניברסיטה'
#sheet['H1'] = 'מידע נוסף'

#sheetG.insert(headline)

sheetG.append(myList)
sheetD.append(myList)
print (sheetD.title)
book.save('appending.xlsx')
